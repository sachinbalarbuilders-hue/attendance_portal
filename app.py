from flask import Flask, render_template, request, jsonify, session, send_from_directory
import pandas as pd
from openpyxl import load_workbook
import hashlib
import datetime
import os
from werkzeug.utils import secure_filename
import tempfile
from database import db
from utils.auth import EmployeeDatabase
from email_service import email_service
import gmail_config  # This will set up Gmail credentials

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Maintenance mode configuration
MAINTENANCE_FLAG_FILE = 'maintenance_mode.flag'

def is_maintenance_mode():
    """Check if maintenance mode is enabled by looking for flag file"""
    return os.path.exists(MAINTENANCE_FLAG_FILE)

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# **FIXED: Single EmployeeDatabase class definition**
class EmployeeDatabase:
    def __init__(self):
        # Initialize with ONLY admin accounts - employees will be auto-created from Excel
        self.EMPLOYEE_DB = {
            # Admin accounts only (updated for Balar Builders)
            "admin@balarbuilders.com": {
                "password": hashlib.sha256("admin123".encode()).hexdigest(),
                "name": "System Administrator",
                "role": "Admin",
                "is_admin": True
            },
            "bhavin@gmail.com": {
                "password": hashlib.sha256("bhavin123".encode()).hexdigest(),
                "name": "Bhavin Patel",
                "role": "Manager",
                "is_admin": True
            }
        }
    
    def clean_employee_name(self, full_name):
        """Clean employee name by removing suffixes like (T), (TC), etc."""
        import re
        cleaned_name = full_name.strip()
        cleaned_name = re.sub(r'\s*\([^)]*\)$', '', cleaned_name)
        return cleaned_name.lower().replace(' ', '').replace('.', '').replace('-', '')
    
    def is_t_employee(self, full_name):
        """Check if employee has (T) suffix - these are not eligible for PL/SL"""
        if not full_name:
            return False
        return '(T)' in full_name.upper()  # Updated: T employees not eligible for PL/SL
    
    def get_employee_leave_eligibility(self, full_name):
        """Get leave eligibility for an employee based on their suffix"""
        if self.is_t_employee(full_name):
            return {
                'W/O': True,   # Week Off - eligible
                'PL': False,   # Personal Leave - NOT eligible
                'SL': False,   # Sick Leave - NOT eligible  
                'FL': True     # Festival Leave - eligible
            }
        else:
            return {
                'W/O': True,   # Week Off - eligible
                'PL': True,    # Personal Leave - eligible
                'SL': True,    # Sick Leave - eligible
                'FL': True     # Festival Leave - eligible
            }

    def create_employee_email(self, full_name):
        """Convert 'Sachin Mandal (T)' to 'sachinmandal@gmail.com'"""
        email_name = self.clean_employee_name(full_name)
        return f"{email_name}@gmail.com"

    def add_employee(self, employee_name):
        """Add new employee to database"""
        email = self.create_employee_email(employee_name)
        password_hash = hashlib.sha256("Balar123".encode()).hexdigest()
        
        if email not in self.EMPLOYEE_DB:
            self.EMPLOYEE_DB[email] = {
                "password": password_hash,
                "name": employee_name,
                "role": "Employee",
                "is_admin": False
            }
            return True, email
        return False, email

    def authenticate_user(self, email, password):
        """Authenticate user credentials"""
        # First check hardcoded admin accounts
        if email in self.EMPLOYEE_DB:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            if self.EMPLOYEE_DB[email]["password"] == hashed_password:
                return True, self.EMPLOYEE_DB[email]
        
        # Then check database for employee accounts
        # Extract employee name from email (reverse of create_employee_email)
        if email.endswith('@gmail.com'):
            email_name = email.replace('@gmail.com', '')
            # Try to find employee in database
            employees = db.get_employees()
            for emp_name in employees:
                # Clean the employee name the same way as in create_employee_email
                cleaned_emp_name = self.clean_employee_name(emp_name)
                
                if email_name == cleaned_emp_name:
                    # Check if user has changed their password before
                    has_ever_changed_password = db.has_user_changed_password(email)
                    
                    if has_ever_changed_password:
                        # User has changed password, only check against stored password in EMPLOYEE_DB
                        if email in self.EMPLOYEE_DB:
                            hashed_password = hashlib.sha256(password.encode()).hexdigest()
                            if self.EMPLOYEE_DB[email]["password"] == hashed_password:
                                return True, {
                                    "password": hashed_password,
                                    "name": emp_name,
                                    "role": "Employee",
                                    "is_admin": False
                                }
                    else:
                        # User still has default password, check against default
                        hashed_password = hashlib.sha256("Balar123".encode()).hexdigest()
                        if hashlib.sha256(password.encode()).hexdigest() == hashed_password:
                            return True, {
                                "password": hashed_password,
                                "name": emp_name,
                                "role": "Employee",
                                "is_admin": False
                            }
        
        return False, None

    def get_all_employees(self):
        """Get list of all employee names (excluding admins)"""
        return [user_data["name"] for email, user_data in self.EMPLOYEE_DB.items()
                if not user_data["is_admin"]]

    def user_exists(self, email):
        """Check if user already exists"""
        # Check hardcoded admin accounts
        if email in self.EMPLOYEE_DB:
            return True
        
        # Check database for employee accounts
        if email.endswith('@gmail.com'):
            email_name = email.replace('@gmail.com', '')
            employees = db.get_employees()
            for emp_name in employees:
                # Clean the employee name the same way as in create_employee_email
                cleaned_emp_name = self.clean_employee_name(emp_name)
                
                if email_name == cleaned_emp_name:
                    return True
        
        return False

    def process_excel_employees(self, unique_employees):
        """Process employee list from Excel and create accounts"""
        created_accounts = []
        existing_accounts = []
        
        for employee_name in unique_employees:
            success, email = self.add_employee(employee_name)
            if success:
                created_accounts.append({
                    'name': employee_name,
                    'email': email,
                    'password': 'Balar123'
                })
            else:
                existing_accounts.append({
                    'name': employee_name,
                    'email': email
                })
        
        return created_accounts, existing_accounts


    def get_user_by_email(self, email):
        """Get user data by email"""
        return self.EMPLOYEE_DB.get(email, None)
    
    def reset_password(self, email, new_password):
        """Reset user password"""
        if email in self.EMPLOYEE_DB:
            self.EMPLOYEE_DB[email]["password"] = hashlib.sha256(new_password.encode()).hexdigest()
            return True
        return False
    
    def has_default_password(self, email):
        """Check if user has default password"""
        if email in self.EMPLOYEE_DB:
            default_password_hash = hashlib.sha256("Balar123".encode()).hexdigest()
            return self.EMPLOYEE_DB[email]["password"] == default_password_hash
        return False

# **FIXED: Single global instance**
employee_db = EmployeeDatabase()

# Test function to demonstrate name cleaning and TC employee detection
def test_name_cleaning():
    """Test the name cleaning functionality and TC employee detection"""
    test_cases = [
        "Sachin Mandal (T)",
        "Priya Sharma (TC)", 
        "Raj Kumar (M)",
        "Anita Singh (T)",
        "Vikram Patel (TC)",
        "Normal Name",
        "Name With (Multiple) (Suffixes)"
    ]
    
    print("Testing Employee Name Cleaning and T Detection:")
    print("=" * 60)
    for name in test_cases:
        cleaned = employee_db.clean_employee_name(name)
        email = employee_db.create_employee_email(name)
        is_t = employee_db.is_t_employee(name)
        eligibility = employee_db.get_employee_leave_eligibility(name)
        print(f"Original: {name}")
        print(f"Cleaned:  {cleaned}")
        print(f"Email:    {email}")
        print(f"T Employee: {is_t}")
        print(f"PL Eligible: {eligibility['PL']}, SL Eligible: {eligibility['SL']}")
        if is_t:
            print(f"Display: PL='Not Eligible', SL='Not Eligible'")
        else:
            print(f"Display: PL=count, SL=count")
        print("-" * 40)

# Uncomment the line below to test the name cleaning
test_name_cleaning()


# Global variables for session-like behavior

def authenticate_user(email, password):
    """Use the centralized authentication"""
    return employee_db.authenticate_user(email, password)

def is_font_red(cell):
    """Check if font color is specifically red or red-like"""
    if cell is None or cell.font is None:
        return False
    
    color = cell.font.color
    if color is None:
        return False
    
    # Handle RGB colors - check for red shades
    if color.type == "rgb" and color.rgb is not None:
        rgb_val = color.rgb.upper()
        red_colors = [
            "FFFF0000", "FF0000", "FFDC143C", "FFB22222", "FF8B0000",
            "FFCD5C5C", "FFF08080", "FFFA8072", "FFFF6347", "FFFF1493"
        ]
        return rgb_val in red_colors
    elif color.type == "indexed" and color.indexed is not None:
        red_indices = [3, 5, 10, 53]
        return color.indexed in red_indices
    elif color.type == "theme" and color.theme is not None:
        return color.theme == 2
    
    return False

def extract_employee_time_range(sheet):
    """Extract time range from employee sheet (e.g., '08:30 AM to 07:00 PM')"""
    try:
        # Look for time range in cell A2 (common location for time range)
        time_cell = sheet.cell(2, 1)  # Row 2, Column A
        if time_cell.value and isinstance(time_cell.value, str):
            time_value = str(time_cell.value).strip()
            # Check if it contains time pattern (AM/PM or 24-hour format)
            if ('AM' in time_value.upper() or 'PM' in time_value.upper() or 
                ':' in time_value or 'to' in time_value.lower()):
                return time_value
        
        # Also check other common locations
        for row in range(1, 6):  # Check first 5 rows
            for col in range(1, 4):  # Check first 3 columns
                cell = sheet.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    time_value = str(cell.value).strip()
                    if ('AM' in time_value.upper() or 'PM' in time_value.upper() or 
                        ':' in time_value or 'to' in time_value.lower()):
                        return time_value
    except Exception:
        pass
    
    return None

def get_ordinal_number(n):
    """Convert number to ordinal (1st, 2nd, 3rd, etc.)"""
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f"{n}{suffix}"

def calculate_late_statistics(employee_name, time_range):
    """Calculate late statistics for an employee based on their time range"""
    try:
        # Parse the time range to get start time
        start_time = None
        if time_range:
            # Extract start time from time range (e.g., "08:30 AM to 07:00 PM" -> "08:30 AM")
            time_parts = time_range.split(' to ')
            if len(time_parts) >= 1:
                start_time_str = time_parts[0].strip()
                try:
                    # Parse different time formats
                    time_formats = [
                        "%I:%M %p",      # 8:30 AM
                        "%H:%M",         # 08:30
                        "%I.%M %p",      # 8.30 AM
                        "%H.%M",         # 08.30
                    ]
                    for fmt in time_formats:
                        try:
                            start_time = datetime.datetime.strptime(start_time_str, fmt).time()
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
        
        # Default to 9:00 AM if no time range found
        if not start_time:
            start_time = datetime.time(9, 0)
        
        # Get attendance records for this employee
        records = db.get_attendance_records(employee_filter=employee_name)
        
        late_count = 0
        total_late_minutes = 0
        late_records = []
        
        for record in records:
            if record.get('Punch-In') and record.get('Date'):
                try:
                    # Parse punch-in time
                    punch_in_str = record['Punch-In']
                    if isinstance(punch_in_str, str):
                        # Handle different time formats
                        time_formats = [
                            "%H:%M:%S",     # 09:30:00
                            "%H:%M",        # 09:30
                            "%I:%M:%S %p",  # 9:30:00 AM
                            "%I:%M %p",     # 9:30 AM
                        ]
                        punch_in_time = None
                        for fmt in time_formats:
                            try:
                                punch_in_time = datetime.datetime.strptime(punch_in_str, fmt).time()
                                break
                            except ValueError:
                                continue
                        
                        if punch_in_time:
                            # Parse the record date
                            record_date = None
                            if isinstance(record['Date'], str):
                                try:
                                    record_date = datetime.datetime.strptime(record['Date'], '%Y-%m-%d').date()
                                except ValueError:
                                    try:
                                        record_date = datetime.datetime.strptime(record['Date'], '%d/%m/%Y').date()
                                    except ValueError:
                                        continue
                            elif hasattr(record['Date'], 'date'):
                                record_date = record['Date'].date()
                            
                            if record_date:
                                # Calculate late minutes (allowing 2 minutes grace period)
                                grace_period = datetime.timedelta(minutes=2)
                                start_datetime = datetime.datetime.combine(record_date, start_time)
                                punch_in_datetime = datetime.datetime.combine(record_date, punch_in_time)
                                
                                late_datetime = punch_in_datetime - start_datetime - grace_period
                                
                                if late_datetime.total_seconds() > 0:
                                    late_minutes = int(late_datetime.total_seconds() / 60)
                                    late_count += 1
                                    total_late_minutes += late_minutes
                                    
                                    # Calculate expected punch-in time
                                    expected_punch_in = start_datetime + grace_period
                                    
                                    late_records.append({
                                        'date': record['Date'],
                                        'punch_in': punch_in_str,
                                        'expected_punch_in': expected_punch_in.strftime("%I:%M %p"),
                                        'late_minutes': late_minutes,
                                        'late_hours': round(late_minutes / 60, 2),
                                        'record_id': f"{employee_name}_{record['Date']}_{punch_in_str}",
                                        'status': record.get('Status', 'P'),
                                        'punch_out': record.get('Punch-Out', ''),
                                        'is_weekend': record_date.weekday() >= 5  # Saturday = 5, Sunday = 6
                                    })
                except Exception as e:
                    print(f"Error processing record for {employee_name}: {e}")
                    continue
        
        # Sort late records by date (oldest first)
        late_records.sort(key=lambda x: x['date'], reverse=False)
        
        # Add sequence numbers to late records
        for i, record in enumerate(late_records):
            record['sequence'] = i + 1
            record['sequence_text'] = get_ordinal_number(i + 1) + " late"
        
        print(f"Late statistics for {employee_name}:")
        print(f"  Total late count: {late_count}")
        print(f"  Total late minutes: {total_late_minutes}")
        print(f"  Late records: {len(late_records)}")
        for i, record in enumerate(late_records[:3]):  # Show first 3 records
            print(f"    Record {i+1}: {record}")
        
        return {
            'total_late_count': late_count,
            'total_late_minutes': total_late_minutes,
            'late_records': late_records,
            'start_time': start_time.strftime("%I:%M %p"),
            'average_late_minutes': round(total_late_minutes / late_count, 2) if late_count > 0 else 0
        }
    except Exception as e:
        print(f"Error calculating late statistics for {employee_name}: {e}")
        return {
            'total_late_count': 0,
            'total_late_minutes': 0,
            'late_records': [],
            'start_time': '09:00 AM'
        }

def process_attendance_file(file_path, selected_date=None):
    """Process attendance data from Excel file"""
    if selected_date is None:
        selected_date = datetime.date.today()
    
    wb = load_workbook(file_path, data_only=False)
    visible_sheets = [sheet for sheet in wb.worksheets if sheet.sheet_state == "visible"]
    
    print(f"DEBUG: Found {len(visible_sheets)} visible sheets in Excel file")
    for sheet in visible_sheets:
        print(f"DEBUG: Sheet name: '{sheet.title}'")
    
    month_abbr = selected_date.strftime("%b").upper()
    year = selected_date.year
    month_num = selected_date.month
    day_limit = selected_date.day
    
    print(f"DEBUG: Processing for month '{month_abbr}' and year '{year}', day limit: {day_limit}")
    print(f"DEBUG: Selected date: {selected_date}")
    
    # Your 5 blank employees
    blank_employees = [
        "Bhavin Patel",
        "Pramod Dubey",
        "Shrikant Talekar",
        "Jitendra Patolia",
        "Lalit Dobariya"
    ]
    
    def to_ts(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return pd.NaT
        
        # Handle datetime.time objects directly
        if isinstance(x, datetime.time):
            return datetime.datetime.combine(datetime.date.today(), x)
        
        # Handle string time formats that might be manually entered
        if isinstance(x, str):
            x = x.strip()
            # Try common time formats first
            time_formats = [
                "%H:%M",      # 09:30
                "%H:%M:%S",   # 09:30:00
                "%I:%M %p",   # 9:30 AM
                "%I:%M:%S %p", # 9:30:00 AM
                "%H.%M",      # 09.30
                "%I.%M %p",   # 9.30 AM
            ]
            
            for fmt in time_formats:
                try:
                    # Parse as time and convert to datetime
                    time_obj = datetime.datetime.strptime(x, fmt).time()
                    # Create a datetime with today's date and the parsed time
                    return datetime.datetime.combine(datetime.date.today(), time_obj)
                except ValueError:
                    continue
        
        try:
            return pd.to_datetime(x, errors="coerce")
        except Exception:
            return pd.NaT
    
    records = []
    
    ignore_statuses = {"A", "W/O", "PL", "SL", "FL", "HL", "PAT", "MAT"}
    # Note: "P", "HF", "PHF", "SHF" are NOT in ignore_statuses, so they will always show punches

    # Extract time ranges for each employee
    employee_time_ranges = {}
    for sheet in visible_sheets:
        time_range = extract_employee_time_range(sheet)
        if time_range:
            employee_time_ranges[sheet.title] = time_range

    processed_sheets = 0
    skipped_sheets = 0
    
    for sheet in visible_sheets:
        df = pd.read_excel(file_path, sheet_name=sheet.title, header=None)
        
        if df.dropna(how="all").empty:
            print(f"DEBUG: Skipping sheet '{sheet.title}' - empty data")
            skipped_sheets += 1
            continue
        
        # Updated logic to handle both old format (JAN, MAY, etc.) and new format (NOV-24, DEC-24, JAN-25, etc.)
        month_rows = []
        print(f"DEBUG: Looking for month '{month_abbr}' in sheet '{sheet.title}'")
        
        for i in range(len(df)):
            cell_value = str(df.iloc[i, 0]).upper()
            print(f"DEBUG: Row {i}: '{cell_value}'")
            
            # Check for old format (JAN, MAY, etc.)
            if cell_value == month_abbr and len(cell_value) == 3:
                print(f"DEBUG: Found old format match '{cell_value}' at row {i}")
                month_rows.append(i)
            # Check for new format (NOV-24, DEC-24, JAN-25, etc.)
            elif cell_value.startswith(month_abbr) and len(cell_value) == 6 and '-' in cell_value and cell_value[3] == '-':
                # Extract year from Excel (e.g., "24" from "NOV-24")
                excel_year_str = cell_value[4:6]
                try:
                    excel_year = int(excel_year_str)
                    # Convert to full year (24 -> 2024, 25 -> 2025, etc.)
                    excel_full_year = 2000 + excel_year
                    print(f"DEBUG: Found new format '{cell_value}', Excel year: {excel_full_year}, Selected year: {year}")
                    
                    # Check if the year matches the selected date year (allow ±1 year flexibility)
                    if excel_full_year == year or excel_full_year == year - 1 or excel_full_year == year + 1:
                        print(f"DEBUG: Year match (flexible)! Adding row {i}")
                        month_rows.append(i)
                    else:
                        print(f"DEBUG: Year too far apart, skipping row {i}")
                except ValueError:
                    # If year parsing fails, skip this row
                    print(f"DEBUG: Invalid year format in '{cell_value}', skipping row {i}")
                    continue
            # Check for date format (like "2025-09-01 00:00:00")
            elif len(cell_value) >= 10 and cell_value.startswith(str(year)) and '-' in cell_value:
                try:
                    # Parse the date to extract month
                    date_part = cell_value.split(' ')[0]  # Get "2025-09-01" part
                    date_obj = datetime.datetime.strptime(date_part, '%Y-%m-%d')
                    if date_obj.month == month_num:
                        print(f"DEBUG: Found date format '{cell_value}', month matches! Adding row {i}")
                        month_rows.append(i)
                    else:
                        print(f"DEBUG: Date format '{cell_value}' month {date_obj.month} doesn't match {month_num}, skipping row {i}")
                except ValueError:
                    print(f"DEBUG: Invalid date format '{cell_value}', skipping row {i}")
                    continue
        
        print(f"DEBUG: Found {len(month_rows)} matching month rows: {month_rows}")
        
        if not month_rows:
            print(f"DEBUG: Skipping sheet '{sheet.title}' - no matching month rows found")
            skipped_sheets += 1
            continue
            
        processed_sheets += 1
        
        i = month_rows[0]
        
        for d in range(1, day_limit + 1):
            col = d + 2
            
            if col >= df.shape[1]:
                continue
            
            date_val = datetime.date(year, selected_date.month, d).strftime('%Y-%m-%d')
            status_row = i + 2
            status = ""
            
            if status_row < len(df):
                s = df.iloc[status_row, col]
                if pd.notna(s):
                    status = str(s).upper().strip()
            
            # Get cells for formatting info
            pin_cell = wb[sheet.title].cell(i + 1, col + 1)
            pout_cell = wb[sheet.title].cell(i + 2, col + 1) if i + 1 < len(df) else None
            status_cell = wb[sheet.title].cell(status_row + 1, col + 1) if status_row < len(df) else None
            
            # Comments logic
            pin_comment, pout_comment, status_comment = "", "", ""
            try:
                if pin_cell and pin_cell.comment:
                    pin_comment = pin_cell.comment.text
                if pout_cell and pout_cell.comment:
                    pout_comment = pout_cell.comment.text
                if status_cell and status_cell.comment:
                    status_comment = status_cell.comment.text
            except:
                pass
            
            # RED font color detection
            pin_highlight = is_font_red(pin_cell) if pin_cell else False
            pout_highlight = is_font_red(pout_cell) if pout_cell else False
            status_highlight = is_font_red(status_cell) if status_cell else False
            
            # **Rules for special statuses**
            if status in ["A", "W/O", "PL", "SL", "FL", "HL", "PAT", "MAT"]:
                # Show no punches for off/leave/paid statuses
                pin, pout = "", ""
            elif sheet.title.strip() in blank_employees:
                pin, pout = "", ""  # Keep blank for exception employees
            else:
                # Read timing data directly from openpyxl cells to preserve data types
                pin_cell_value = pin_cell.value if pin_cell else None
                pout_cell_value = pout_cell.value if pout_cell else None
                
                # Also try pandas data as fallback
                raw1 = df.iloc[i, col] if pd.notna(df.iloc[i, col]) else None
                raw2 = df.iloc[i + 1, col] if i + 1 < len(df) and pd.notna(df.iloc[i + 1, col]) else None
                
                # Use openpyxl values first, fallback to pandas
                t1 = to_ts(pin_cell_value) if pin_cell_value is not None else to_ts(raw1)
                t2 = to_ts(pout_cell_value) if pout_cell_value is not None else to_ts(raw2)
                
                if pd.notna(t1) and pd.notna(t2):
                    pin, pout = t1.strftime("%H:%M"), t2.strftime("%H:%M")
                    if status in ["HF", "PHF", "SHF"]:
                        print(f"DEBUG: {status} status with both times: {sheet.title} on {date_val} - {pin} to {pout}")
                elif pd.notna(t1) and pd.isna(t2):
                    # Show punches for P, HF, SHF, PHF statuses even when one is missing
                    if status in ["P", "HF", "PHF", "SHF"]:
                        if t1.hour >= 12:
                            pin, pout = "⚠️ MISSING", t1.strftime("%H:%M")
                        else:
                            pin, pout = t1.strftime("%H:%M"), "⚠️ MISSING"
                        if status in ["HF", "PHF", "SHF"]:
                            print(f"DEBUG: {status} status with missing time: {sheet.title} on {date_val} - {pin} to {pout}")
                    elif status not in ignore_statuses:
                        if t1.hour >= 12:
                            pin, pout = "⚠️ MISSING", t1.strftime("%H:%M")
                        else:
                            pin, pout = t1.strftime("%H:%M"), "⚠️ MISSING"
                        if status in ["HF", "PHF", "SHF"]:
                            print(f"DEBUG: {status} status with missing time: {sheet.title} on {date_val} - {pin} to {pout}")
                    else:
                        # Ignore missing, hide punches for leave/off
                        pin, pout = "", ""
                elif pd.isna(t1) and pd.notna(t2):
                    # Show punches for P, HF, SHF, PHF statuses even when one is missing
                    if status in ["P", "HF", "PHF", "SHF"]:
                        if t2.hour >= 12:
                            pin, pout = "⚠️ MISSING", t2.strftime("%H:%M")
                        else:
                            pin, pout = t2.strftime("%H:%M"), "⚠️ MISSING"
                        if status in ["HF", "PHF", "SHF"]:
                            print(f"DEBUG: {status} status with missing time: {sheet.title} on {date_val} - {pin} to {pout}")
                    elif status not in ignore_statuses:
                        if t2.hour >= 12:
                            pin, pout = "⚠️ MISSING", t2.strftime("%H:%M")
                        else:
                            pin, pout = t2.strftime("%H:%M"), "⚠️ MISSING"
                        if status in ["HF", "PHF", "SHF"]:
                            print(f"DEBUG: {status} status with missing time: {sheet.title} on {date_val} - {pin} to {pout}")
                    else:
                        pin, pout = "", ""
                else:
                    # Both punches missing
                    # Show punches for P, HF, SHF, PHF statuses even when both are missing
                    if status in ["P", "HF", "PHF", "SHF"]:
                        pin, pout = "⚠️ MISSING", "⚠️ MISSING"
                        if status in ["HF", "PHF", "SHF"]:
                            print(f"DEBUG: {status} status with both times missing: {sheet.title} on {date_val}")
                    elif status not in ignore_statuses:
                        pin, pout = "⚠️ MISSING", "⚠️ MISSING"
                        if status in ["HF", "PHF", "SHF"]:
                            print(f"DEBUG: {status} status with both times missing: {sheet.title} on {date_val}")
                    else:
                        pin, pout = "", ""


            records.append({
                "Employee": sheet.title,
                "Date": date_val,
                "Punch-In": pin,
                "Punch-Out": pout,
                "Status": status,
                "pin_comment": pin_comment,
                "pout_comment": pout_comment,
                "status_comment": status_comment,
                "pin_highlight": pin_highlight,
                "pout_highlight": pout_highlight,
                "status_highlight": status_highlight,
                "time_range": employee_time_ranges.get(sheet.title, "")
            })
    
    print(f"DEBUG: Generated {len(records)} total records from all sheets")
    print(f"DEBUG: Processed {processed_sheets} sheets, skipped {skipped_sheets} sheets")
    return records


def extract_leave_totals(file_path):
    """Parse cumulative W/O, PL, SL, FL totals per employee sheet from the Excel.
    Strategy: find header row containing these labels, then take the last numeric value
    in each corresponding column as the sheet's cumulative total.
    TC employees are not eligible for PL/SL.
    """
    wb = load_workbook(file_path, data_only=True)
    visible_sheets = [sheet for sheet in wb.worksheets if sheet.sheet_state == "visible"]

    per_employee = {}

    target_labels = {
        "W/O": ["W/O", "W O", "W-0", "W-O"],
        "PL": ["PL"],
        "SL": ["SL"],
        "FL": ["FL"],
    }

    for sheet in visible_sheets:
        ws = wb[sheet.title]
        employee_name = sheet.title
        
        # Check if this employee is T (not eligible for PL/SL)
        is_t_employee = employee_db.is_t_employee(employee_name)

        # Search header rows (first 10 rows) for our labels
        label_to_col = {}
        max_header_rows = min(10, ws.max_row)
        for r in range(1, max_header_rows + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if not isinstance(val, str):
                    continue
                text = val.strip().upper()
                for key, variants in target_labels.items():
                    if text in [v.upper() for v in variants]:
                        label_to_col[key] = c
            # Small optimization: if all found, stop searching
            if len(label_to_col) == len(target_labels):
                break

        if not label_to_col:
            continue

        # Walk from bottom up to locate last numeric value for each label
        totals = {"W/O": 0, "PL": 0, "SL": 0, "FL": 0}
        for key, col in label_to_col.items():
            for r in range(ws.max_row, 0, -1):
                cell = ws.cell(row=r, column=col)
                val = cell.value
                if isinstance(val, (int, float)):
                    try:
                        totals[key] = float(val)
                    except Exception:
                        totals[key] = 0
                    break

        # For T employees, set PL and SL to "FL" (Festival Leave)
        if is_t_employee:
            totals["PL"] = "FL"
            totals["SL"] = "FL"
            print(f"T Employee '{employee_name}' - PL/SL set to 'FL'")

        per_employee[sheet.title] = totals

    return per_employee

# Routes
@app.route('/')
def index():
    if is_maintenance_mode():
        return render_template('maintenance.html')
    return render_template('index.html')

@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json', mimetype='application/json')


@app.route('/maintenance')
def maintenance():
    return render_template('maintenance.html')

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email')
    password = data.get('password')

    if not email or not password:
        return jsonify({'success': False, 'message': 'Email and password required'})

    is_valid, user_data = authenticate_user(email, password)

    if is_valid:
        session['user_email'] = email
        session['user_data'] = user_data
        
        # Log the login
        ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'Unknown'))
        user_agent = request.headers.get('User-Agent', 'Unknown')
        db.log_login(email, user_data.get('name', 'Unknown'), user_data.get('is_admin', False), ip_address, user_agent)
        
        # Check if user has default password AND has never changed password before
        has_default_password = employee_db.has_default_password(email)
        has_ever_changed_password = db.has_user_changed_password(email)
        needs_password_change = has_default_password and not has_ever_changed_password
        
        print(f"DEBUG: Password change check for {email}:")
        print(f"  - has_default_password: {has_default_password}")
        print(f"  - has_ever_changed_password: {has_ever_changed_password}")
        print(f"  - needs_password_change: {needs_password_change}")
        
        return jsonify({
            'success': True, 
            'user': user_data,
            'needs_password_change': needs_password_change,
            'leave_notification': not user_data.get('is_admin', False)  # Only show for non-admin users
        })

    return jsonify({'success': False, 'message': 'Invalid credentials'})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/send-otp', methods=['POST'])
def send_otp():
    """Send OTP to user's email for password change verification"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
    data = request.json
    actual_email = data.get('actual_email')
    
    if not actual_email:
        return jsonify({'success': False, 'message': 'Email address is required'})
    
    # Validate email format
    if '@' not in actual_email or '.' not in actual_email.split('@')[1]:
        return jsonify({'success': False, 'message': 'Please enter a valid email address'})
    
    email = session['user_email']
    user_data = session['user_data']
    
    try:
        # Generate OTP
        otp_code = email_service.generate_otp()
        
        # Store OTP in database
        if db.store_otp(email, otp_code, actual_email):
            # Send OTP email
            if email_service.send_otp_email(actual_email, otp_code, user_data.get('name', 'Employee')):
                return jsonify({
                    'success': True, 
                    'message': f'OTP sent to {actual_email}',
                    'email': actual_email
                })
            else:
                return jsonify({'success': False, 'message': 'Failed to send OTP email'})
        else:
            return jsonify({'success': False, 'message': 'Failed to generate OTP'})
            
    except Exception as e:
        print(f"Error sending OTP: {e}")
        return jsonify({'success': False, 'message': 'Error sending OTP'})

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    """Verify OTP and change password"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
    data = request.json
    otp_code = data.get('otp_code')
    new_password = data.get('new_password')
    
    if not otp_code:
        return jsonify({'success': False, 'message': 'OTP code is required'})
    
    if not new_password:
        return jsonify({'success': False, 'message': 'New password is required'})
    
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'})
    
    email = session['user_email']
    user_data = session['user_data']
    
    try:
        # Verify OTP
        if db.verify_otp(email, otp_code):
            # Get the actual email from OTP record
            actual_email = db.get_actual_email(email)
            
            # Change password
            if employee_db.reset_password(email, new_password):
                # Log password change for admin visibility
                db.log_password_change(
                    email=email,
                    employee_name=user_data.get('name', 'Unknown'),
                    current_password=new_password,
                    changed_by='employee'
                )
                # Mark that user has changed their password and store actual email
                db.mark_password_as_changed(email, actual_email)
                
                # Clean up expired OTPs
                db.cleanup_expired_otps()
                
                return jsonify({'success': True, 'message': 'Password changed successfully'})
            else:
                return jsonify({'success': False, 'message': 'Failed to change password'})
        else:
            return jsonify({'success': False, 'message': 'Invalid or expired OTP'})
            
    except Exception as e:
        print(f"Error verifying OTP: {e}")
        return jsonify({'success': False, 'message': 'Error verifying OTP'})

@app.route('/api/get-stored-email')
def get_stored_email():
    """Get the stored email address for the current user"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
    email = session['user_email']
    actual_email = db.get_actual_email(email)
    
    return jsonify({
        'success': True,
        'has_stored_email': actual_email is not None,
        'email': actual_email
    })

@app.route('/api/change-password', methods=['POST'])
def change_password():
    """Legacy change password endpoint - kept for backward compatibility"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
    data = request.json
    new_password = data.get('new_password')
    
    if not new_password:
        return jsonify({'success': False, 'message': 'New password is required'})
    
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'})
    
    email = session['user_email']
    user_data = session['user_data']
    
    # Change password
    if employee_db.reset_password(email, new_password):
        # Log password change for admin visibility
        db.log_password_change(
            email=email,
            employee_name=user_data.get('name', 'Unknown'),
            current_password=new_password,
            changed_by='employee'
        )
        # Mark that user has changed their password
        db.mark_password_as_changed(email)
        return jsonify({'success': True, 'message': 'Password changed successfully'})
    
    return jsonify({'success': False, 'message': 'Failed to change password'})

@app.route('/api/login-logs')
def get_login_logs():
    """Get login logs (admin only)"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    if not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})
    
    try:
        logs = db.get_login_logs(limit=200)  # Get last 200 logins
        return jsonify({'success': True, 'logs': logs})
    except Exception as e:
        print(f"Error getting login logs: {e}")
        return jsonify({'success': False, 'message': 'Error retrieving login logs'})

@app.route('/api/current-password')
def get_current_password():
    """Get current password for an employee (admin only)"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    if not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})
    
    email = request.args.get('email')
    if not email:
        return jsonify({'success': False, 'message': 'Email parameter required'})
    
    # Get user data
    user_data = employee_db.get_user_by_email(email)
    print(f"DEBUG: User data for {email}: {user_data}")
    
    if not user_data:
        # Let's also check what emails are available
        all_employees = employee_db.get_all_employees()
        all_emails = employee_db.get_all_employee_emails()
        print(f"DEBUG: Available employees: {all_employees}")
        print(f"DEBUG: Available emails: {all_emails}")
        return jsonify({'success': False, 'message': f'Employee not found for email: {email}'})
    
    # Get the current password from password history (most recent)
    history = db.get_password_history(limit=1000)  # Get all history to find this employee
    current_password = "Balar123"  # Default password
    
    print(f"DEBUG: Looking for password for email: {email}")
    print(f"DEBUG: Password history entries: {len(history)}")
    
    # Find the most recent password for this employee
    for entry in history:
        print(f"DEBUG: Checking entry for {entry['email']}")
        if entry['email'] == email:
            current_password = entry['current_password']
            print(f"DEBUG: Found password: {current_password}")
            break
    
    print(f"DEBUG: Final password: {current_password}")
    
    return jsonify({
        'success': True, 
        'password': current_password,
        'employee_name': user_data.get('name', 'Unknown')
    })



@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})

    # Support multiple files under key 'files'
    files = request.files.getlist('files')
    if not files:
        return jsonify({'success': False, 'message': 'No files uploaded'})

    # Filter valid Excel files
    valid_files = [f for f in files if f and f.filename and f.filename.lower().endswith('.xlsx')]
    if not valid_files:
        return jsonify({'success': False, 'message': 'No valid .xlsx files selected'})

    try:
        selected_date = datetime.date.today()
        if 'selected_date' in request.form:
            selected_date = datetime.datetime.strptime(request.form['selected_date'], '%Y-%m-%d').date()

        total_records = 0
        uploaded_filepaths = []
        created_accounts = []
        existing_accounts = []

        for file in valid_files:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            uploaded_filepaths.append(filepath)

            # Process attendance data
            file_records = process_attendance_file(filepath, selected_date)
            print(f"DEBUG: Generated {len(file_records)} records from file processing")
            
            # Save to database (this will overwrite existing data for this file)
            records_saved = db.save_attendance_records(file_records, filename)
            print(f"DEBUG: Successfully saved {records_saved} records to database")
            total_records += records_saved

            # Process and save leave totals
            sheet_totals = extract_leave_totals(filepath)
            db.save_leave_totals(sheet_totals, filename)

            # Auto-create employee accounts from this file's data
            unique_employees = list(set([record['Employee'] for record in file_records]))
            print(f"DEBUG: Found {len(unique_employees)} unique employees in file: {unique_employees}")
            file_created, file_existing = employee_db.process_excel_employees(unique_employees)
            print(f"DEBUG: Created {len(file_created)} new accounts, {len(file_existing)} existing accounts")
            created_accounts.extend(file_created)
            existing_accounts.extend(file_existing)

        # Cleanup uploaded files
        for p in uploaded_filepaths:
            if os.path.exists(p):
                os.remove(p)

        message = f"Processed {len(valid_files)} file(s), {total_records} total records saved to database. "
        if created_accounts:
            message += f"{len(created_accounts)} new employee accounts created."

        return jsonify({
            'success': True,
            'message': message,
            'record_count': total_records,
            'files_processed': len(valid_files),
            'created_accounts': created_accounts,
            'total_employees': len(set([acc['name'] for acc in created_accounts + existing_accounts])),
            'new_accounts': len(created_accounts)
        })

    except Exception as e:
        # Cleanup on error
        for p in uploaded_filepaths:
            if os.path.exists(p):
                os.remove(p)
        return jsonify({'success': False, 'message': f'Error processing files: {str(e)}'})

@app.route('/api/attendance')
def get_attendance():
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})

    user_data = session['user_data']
    filter_status = request.args.get('status', 'All')
    employee_filter = request.args.get('employee')

    # Determine which employee to filter by
    if not user_data.get('is_admin') or employee_filter:
        target_employee = employee_filter if employee_filter else user_data['name']
    else:
        target_employee = None

    # Get data from database
    filtered_data = db.get_attendance_records(target_employee, filter_status)
    
    return jsonify({'success': True, 'data': filtered_data})

@app.route('/api/employees')
def get_employees():
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})

    employees = db.get_employees()
    return jsonify({'success': True, 'employees': sorted(employees)})

@app.route('/api/leave-totals')
def get_leave_totals():
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})

    user_data = session['user_data']
    employee = request.args.get('employee')

    # Non-admins can only view their own totals
    if not user_data.get('is_admin'):
        employee = user_data['name']

    # Get data from database
    totals = db.get_leave_totals(employee)
    
    if employee and employee not in totals:
        # Check if this employee is T to set appropriate defaults
        if employee_db.is_t_employee(employee):
            totals[employee] = {"W/O": 0, "PL": "FL", "SL": "FL", "FL": 0}
        else:
            totals[employee] = {"W/O": 0, "PL": 0, "SL": 0, "FL": 0}

    return jsonify({'success': True, 'data': totals})

@app.route('/api/employee-leave-eligibility')
def get_employee_leave_eligibility():
    """Get leave eligibility for an employee (T employees not eligible for PL/SL)"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})

    user_data = session['user_data']
    employee = request.args.get('employee')

    # Non-admins can only view their own eligibility
    if not user_data.get('is_admin'):
        employee = user_data['name']

    if not employee:
        return jsonify({'success': False, 'message': 'Employee name required'})

    # Get leave eligibility based on employee suffix
    eligibility = employee_db.get_employee_leave_eligibility(employee)
    
    return jsonify({
        'success': True, 
        'employee': employee,
        'is_t_employee': employee_db.is_t_employee(employee),
        'eligibility': eligibility
    })

@app.route('/api/admin/set-date', methods=['POST'])
def set_admin_date():
    """Set admin's selected date"""
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})
    
    data = request.get_json()
    selected_date = data.get('date')
    
    if not selected_date:
        return jsonify({'success': False, 'message': 'Date is required'})
    
    success = db.set_admin_setting('selected_date', selected_date)
    
    if success:
        return jsonify({'success': True, 'message': 'Date updated successfully'})
    else:
        return jsonify({'success': False, 'message': 'Failed to update date'})

@app.route('/api/admin/get-date')
def get_admin_date():
    """Get admin's selected date"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    selected_date = db.get_admin_setting('selected_date')
    
    return jsonify({
        'success': True, 
        'date': selected_date,
        'has_date': selected_date is not None
    })

@app.route('/api/database-stats')
def get_database_stats():
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})

    stats = db.get_database_stats()
    return jsonify({'success': True, 'stats': stats})

@app.route('/api/upload-history')
def get_upload_history():
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})

    history = db.get_upload_history()
    return jsonify({'success': True, 'history': history})

@app.route('/api/clear-database', methods=['POST'])
def clear_database():
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})

    try:
        db.clear_existing_data()
        return jsonify({'success': True, 'message': 'Database cleared successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error clearing database: {str(e)}'})










@app.route('/api/late-statistics')
def get_late_statistics():
    """Get late statistics for the current user"""
    if 'user_data' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
    user_data = session['user_data']
    employee_name = user_data['name']
    
    print(f"Calculating late statistics for: {employee_name}")
    
    # Get the time range for this employee from their records
    records = db.get_attendance_records(employee_filter=employee_name)
    print(f"Found {len(records)} records for {employee_name}")
    
    time_range = None
    if records:
        # Get the most recent time range from records
        for record in records:
            if record.get('time_range'):
                time_range = record['time_range']
                print(f"Found time range: {time_range}")
                break
    
    if not time_range:
        print("No time range found, using default 9:00 AM")
    
    # Calculate late statistics
    late_stats = calculate_late_statistics(employee_name, time_range)
    print(f"Late stats result: {late_stats}")
    
    return jsonify({
        'success': True,
        'data': late_stats
    })

@app.route('/api/admin/late-statistics')
def get_admin_late_statistics():
    """Get late statistics for all employees (admin only)"""
    print(f"Admin late statistics request - User: {session.get('user_data', {}).get('name', 'Unknown')}")
    
    if 'user_data' not in session or not session['user_data'].get('is_admin'):
        return jsonify({'success': False, 'message': 'Admin access required'})
    
    try:
        # Get all attendance records
        all_records = db.get_attendance_records()
        
        # Group records by employee
        employee_records = {}
        for record in all_records:
            employee_name = record['Employee']
            if employee_name not in employee_records:
                employee_records[employee_name] = []
            employee_records[employee_name].append(record)
        
        # Calculate late statistics for each employee
        all_employee_stats = []
        
        for employee_name, records in employee_records.items():
            # Get time range for this employee
            time_range = None
            for record in records:
                if record.get('time_range'):
                    time_range = record['time_range']
                    break
            
            # Calculate late statistics
            late_stats = calculate_late_statistics(employee_name, time_range)
            
            # Add employee name to the stats
            late_stats['employee_name'] = employee_name
            all_employee_stats.append(late_stats)
        
        # Sort by total late count (descending)
        all_employee_stats.sort(key=lambda x: x['total_late_count'], reverse=True)
        
        print(f"Admin late statistics - Total employees: {len(all_employee_stats)}, Employees with late arrivals: {len([e for e in all_employee_stats if e['total_late_count'] > 0])}")
        
        return jsonify({
            'success': True,
            'data': all_employee_stats
        })
        
    except Exception as e:
        print(f"Error calculating admin late statistics: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to calculate late statistics'
        })


if __name__ == '__main__':
    app.run(debug=True)
